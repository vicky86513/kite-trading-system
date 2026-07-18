"""
Shared price-book Excel writer used by every live tracker.

Keeps an in-memory hashmap: price -> {bid_qty, ask_qty, last_updated}.
- No duplicate strike-price rows ever get written.
- Every time a price shows up again, its quantity is ADDED on top of what's
  already stored there (running total at that price), and "last updated" is
  refreshed.
- The Excel sheet is fully rewritten from the hashmap on every flush(),
  sorted ascending by price.
- All output (.xlsx, .json, .html) is written into the Excel/ subfolder,
  so the main folder stays clean. The folder is created automatically if
  it doesn't exist.
"""

import os
import math
import json as _json
import threading
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from . import config

OUTPUT_DIR = config.OUTPUT_DIR
PRICE_ROUND_STEP = 10


def round_price(value, step=PRICE_ROUND_STEP):
    """Rounds `value` to the nearest multiple of `step`.
    
    Threshold is step/2: strictly more than that goes UP to the next
    bucket, otherwise it goes DOWN to the current one.
    
    e.g. with step=10:
        1002 -> 1000 (remainder 2, <=5, rounds down)
        1006 -> 1010 (remainder 6, >5, rounds up)
    
    Pass step=None to disable rounding entirely — the exact price is used
    as-is (still None-safe).
    """
    if value is None:
        return None
    if step is None:
        return value

    base = math.floor(value / step) * step
    remainder = value - base
    threshold = step / 2
    bucketed = base + step if remainder > threshold else base
    return int(round(bucketed))


class PriceBookWriter:
    """Writes live price book data to Excel/JSON/HTML."""

    HEADERS = [
        "Strike Price",
        "Bid Qty (cum.)",
        "Ask Qty (cum.)",
        "Net (Bid-Ask)",
        "Total Qty (Bid+Ask)",
        "Last Updated",
    ]

    @staticmethod
    def _dated_filepath(filename):
        """Inserts today's date before the extension.
        
        e.g. 'crude_oil_price_book.xlsx' ->
             'data/Excel/crude_oil_price_book_2026-07-17.xlsx'
        """
        name, ext = os.path.splitext(filename)
        date_str = datetime.now().strftime("%Y-%m-%d")
        return OUTPUT_DIR / f"{name}_{date_str}{ext}"

    def __init__(self, filename, title="Live Price Book", price_step=PRICE_ROUND_STEP):
        """`filename` is just a plain filename like 'crude_oil_price_book.xlsx' —
        it will be created inside the Excel/ subfolder automatically, with
        today's date stamped into the name (e.g.
        'crude_oil_price_book_2026-07-17.xlsx'), so every day starts a brand
        new file instead of accumulating forever.

        If the script happens to stay running past midnight, flush()
        automatically rolls over to a fresh file (and a fresh, empty price
        book) for the new day.

        `price_step` controls the price bucketing (see round_price above) —
        defaults to 10 (futures), pass price_step=None for options.
        """
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        self._original_filename = filename
        self.title = title
        self.price_step = price_step
        self.lock = threading.Lock()
        self.book = {}  # price -> {"bid_qty": int, "ask_qty": int, "last_updated": str}
        self.current_date = datetime.now().date()

        filepath = self._dated_filepath(filename)
        self.filepath = filepath
        self.html_filepath = filepath.with_suffix(".html")
        self.json_filepath = filepath.with_suffix(".json")

        # Try to load existing file
        if filepath.exists():
            try:
                wb = load_workbook(filepath)
                ws = wb.active

                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                if header != self.HEADERS:
                    # Different format — back it up
                    backup_path = filepath.with_name(f"{filepath.stem}_old_format_backup{filepath.suffix}")
                    if backup_path.exists():
                        backup_path.unlink()
                    filepath.rename(backup_path)
                    print(
                        f"Existing {filepath} was a different format — "
                        f"backed it up to {backup_path} and starting fresh."
                    )
                else:
                    # Rehydrate hashmap so restarting doesn't wipe out today's data
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or row[0] is None:
                            continue
                        price, bid_qty, ask_qty, _net, _total, last_updated = row
                        price = round_price(price, step=self.price_step)
                        entry = self.book.setdefault(
                            price, {"bid_qty": 0, "ask_qty": 0, "last_updated": None}
                        )
                        entry["bid_qty"] += bid_qty or 0
                        entry["ask_qty"] += ask_qty or 0
                        # keep whichever timestamp is more recent
                        if last_updated and (
                            entry["last_updated"] is None or last_updated > entry["last_updated"]
                        ):
                            entry["last_updated"] = last_updated
            except Exception as e:
                print(f"Warning: Could not load existing file {filepath}: {e}")

    def _check_daily_rollover(self):
        """If the calendar date has changed since this writer was created
        (i.e. the script has been running across midnight), switch to a
        fresh dated filename and start today's price book empty.

        Cheap to call often — it's just a date comparison.
        """
        today = datetime.now().date()
        if today != self.current_date:
            self.current_date = today
            self.filepath = self._dated_filepath(self._original_filename)
            self.html_filepath = self.filepath.with_suffix(".html")
            self.json_filepath = self.filepath.with_suffix(".json")
            self.book = {}
            print(f"📅 New day detected — starting fresh price book: {self.filepath}")

    def update(self, price, bid_qty=None, ask_qty=None):
        """Add bid_qty and/or ask_qty onto whatever is already stored at this price.

        Price is rounded first using this writer's price_step (>.5*step rounds
        up, <=.5*step rounds down), e.g. with price_step=10: 5820.3 and 5820.4
        both land in the same 5820 row.
        """
        self._check_daily_rollover()
        price = round_price(price, step=self.price_step)
        if price is None:
            return

        with self.lock:
            entry = self.book.setdefault(price, {"bid_qty": 0, "ask_qty": 0, "last_updated": None})
            if bid_qty:
                entry["bid_qty"] += bid_qty
            if ask_qty:
                entry["ask_qty"] += ask_qty
            entry["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def flush(self):
        """Write accumulated price book to Excel/JSON/HTML."""
        self._check_daily_rollover()

        with self.lock:
            # Make snapshot so updates don't affect this flush
            book_snapshot = {p: dict(e) for p, e in self.book.items()}

        # Write Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "PriceBook"
        ws.append(self.HEADERS)

        for price in sorted(book_snapshot.keys()):
            entry = book_snapshot[price]
            net = entry["bid_qty"] - entry["ask_qty"]
            total_qty = entry["bid_qty"] + entry["ask_qty"]
            ws.append(
                [price, entry["bid_qty"], entry["ask_qty"], net, total_qty, entry["last_updated"]]
            )

        self.filepath.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.filepath)

        # Write JSON & HTML
        self._write_json(book_snapshot)
        self._write_html(book_snapshot)

    def _write_json(self, book_snapshot):
        """Write price book as JSON for dashboard."""
        rows = []
        for price in sorted(book_snapshot.keys()):
            entry = book_snapshot[price]
            rows.append(
                {
                    "price": price,
                    "bid_qty": entry["bid_qty"],
                    "ask_qty": entry["ask_qty"],
                    "net": entry["bid_qty"] - entry["ask_qty"],
                    "total_qty": entry["bid_qty"] + entry["ask_qty"],
                    "last_updated": entry["last_updated"],
                }
            )

        payload = {
            "title": self.title,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "rows": rows,
        }

        with open(self.json_filepath, "w", encoding="utf-8") as f:
            _json.dump(payload, f, indent=2)

    def _write_html(self, book_snapshot):
        """Write price book as HTML for quick viewing."""
        prices = sorted(book_snapshot.keys())
        now = datetime.now()

        total_bid = sum(e["bid_qty"] for e in book_snapshot.values())
        total_ask = sum(e["ask_qty"] for e in book_snapshot.values())
        total_net = total_bid - total_ask
        max_abs_net = (
            max((abs(e["bid_qty"] - e["ask_qty"]) for e in book_snapshot.values()), default=0)
            or 1
        )

        rows_html = []
        for price in prices:
            entry = book_snapshot[price]
            bid_qty = entry["bid_qty"]
            ask_qty = entry["ask_qty"]
            net = bid_qty - ask_qty

            if net > 0:
                net_class, net_sign, bar_side = "net-positive", "+", "bar-buy"
            elif net < 0:
                net_class, net_sign, bar_side = "net-negative", "", "bar-sell"
            else:
                net_class, net_sign, bar_side = "net-neutral", "", ""

            bar_pct = round(abs(net) / max_abs_net * 100, 1)

            # Freshness glow — rows touched in the last 10s get a live pulse
            fresh_class = ""
            if entry["last_updated"]:
                try:
                    updated_dt = datetime.strptime(entry["last_updated"], "%Y-%m-%d %H:%M:%S")
                    if (now - updated_dt).total_seconds() <= 10:
                        fresh_class = " fresh"
                except ValueError:
                    pass

            rows_html.append(
                f"""
                
                    {price}
                    {bid_qty:,}
                    {ask_qty:,}
                    
                        
                            
                            
                        
                        {net_sign}{net:,}
                    
                    {bid_qty + ask_qty:,}
                    {entry['last_updated'] or '-'}
                """
            )

        rows_joined = "".join(rows_html) if rows_html else (
            'Waiting for live ticks…'
        )

        net_summary_class = (
            "net-positive" if total_net > 0 else ("net-negative" if total_net < 0 else "net-neutral")
        )
        net_summary_sign = "+" if total_net > 0 else ""

        html = f"""




{self.title} · Price Book



    :root {{
        --bg: #08090c;
        --bg-glow: #10141d;
        --panel: #11151d;
        --panel-2: #151a24;
        --border: #232a38;
        --text: #e8ecf1;
        --muted: #6f7a8a;
        --buy: #00d38a;
        --sell: #ff4d6a;
        --amber: #f5b53f;
    }}
    * {{ box-sizing: border-box; }}
    body {{
        margin: 0;
        font-family: 'IBM Plex Mono', ui-monospace, monospace;
        background:
            radial-gradient(1100px 500px at 12% -10%, var(--bg-glow), transparent 60%),
            var(--bg);
        color: var(--text);
        padding: 44px 24px 60px;
    }}
    .wrap {{ max-width: 980px; margin: 0 auto; }}

    header {{
        display: flex;
        justify-content: space-between;
        align-items: flex-end;
        flex-wrap: wrap;
        gap: 14px;
        margin-bottom: 26px;
    }}
    .title-block h1 {{
        font-family: 'Space Grotesk', sans-serif;
        font-size: 26px;
        font-weight: 700;
        letter-spacing: 0.2px;
        margin: 0 0 4px;
    }}
    .title-block .eyebrow {{
        display: flex;
        align-items: center;
        gap: 8px;
        color: var(--amber);
        font-size: 11.5px;
        text-transform: uppercase;
        letter-spacing: 1.5px;
        margin-bottom: 8px;
    }}
    .live-dot {{
        width: 7px; height: 7px; border-radius: 50%;
        background: var(--amber);
        animation: pulse 1.8s infinite;
    }}
    @keyframes pulse {{
        0%   {{ box-shadow: 0 0 0 0 rgba(245,181,63,0.55); }}
        70%  {{ box-shadow: 0 0 0 8px rgba(245,181,63,0); }}
        100% {{ box-shadow: 0 0 0 0 rgba(245,181,63,0); }}
    }}
    .subtitle {{ color: var(--muted); font-size: 12.5px; }}

    .stats {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(120px, 1fr));
        gap: 10px;
        margin-bottom: 20px;
    }}
    .stat {{
        background: var(--panel);
        border: 1px solid var(--border);
        border-radius: 10px;
        padding: 14px 16px;
    }}
    .stat .label {{
        font-family: 'Space Grotesk', sans-serif;
        color: var(--muted);
        font-size: 10.5px;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-bottom: 6px;
    }}
    .stat .value {{ font-size: 18px; font-weight: 600; font-variant-numeric: tabular-nums; }}
    .stat .value.buy {{ color: var(--buy); }}
    .stat .value.sell {{ color: var(--sell); }}

    .panel {{
        background: var(--panel);
        border: 1px solid var(--border);
        border-radius: 14px;
        overflow: hidden;
        box-shadow: 0 20px 50px rgba(0,0,0,0.4);
    }}
    table {{ width: 100%; border-collapse: collapse; }}
    thead th {{
        text-align: right;
        font-family: 'Space Grotesk', sans-serif;
        font-size: 11px;
        text-transform: uppercase;
        letter-spacing: 0.8px;
        color: var(--muted);
        padding: 14px 18px;
        border-bottom: 1px solid var(--border);
        background: var(--panel-2);
    }}
    thead th:first-child {{ text-align: left; }}
    tbody td {{
        text-align: right;
        padding: 13px 18px;
        border-bottom: 1px solid rgba(255,255,255,0.035);
        font-variant-numeric: tabular-nums;
        font-size: 13.5px;
        vertical-align: middle;
    }}
    tbody tr.row {{ transition: background 0.2s; }}
    tbody tr.row:hover {{ background: rgba(255,255,255,0.02); }}
    tbody tr.fresh {{ background: rgba(245,181,63,0.05); }}
    tbody tr.fresh td.price {{ position: relative; }}
    tbody tr.fresh td.price::before {{
        content: "";
        position: absolute; left: -10px; top: 50%;
        width: 5px; height: 5px; border-radius: 50%;
        background: var(--amber);
        transform: translateY(-50%);
        animation: pulse 1.8s infinite;
    }}
    td.price {{ text-align: left; font-weight: 600; color: var(--text); padding-left: 26px; }}
    td.bid {{ color: var(--buy); }}
    td.ask {{ color: var(--sell); }}
    td.net-cell {{ min-width: 190px; }}
    td.net-cell > span:last-child {{ display: inline-block; min-width: 64px; font-weight: 600; }}
    .net-positive {{ color: var(--buy); }}
    .net-negative {{ color: var(--sell); }}
    .net-neutral {{ color: var(--muted); }}
    .imbalance-track {{
        position: relative;
        display: inline-block;
        width: 80px;
        height: 6px;
        border-radius: 3px;
        background: rgba(255,255,255,0.06);
        vertical-align: middle;
        margin-right: 10px;
        overflow: hidden;
    }}
    .imbalance-mid {{
        position: absolute; left: 50%; top: -2px;
        width: 1px; height: 10px;
        background: rgba(255,255,255,0.15);
    }}
    .imbalance-fill {{
        position: absolute; top: 0; height: 100%;
        border-radius: 3px;
    }}
    .imbalance-fill.bar-buy {{ left: 50%; background: var(--buy); }}
    .imbalance-fill.bar-sell {{ right: 50%; background: var(--sell); }}
    td.time {{ color: var(--muted); font-size: 11.5px; }}
    td.total-qty {{ color: var(--text); font-weight: 600; }}
    td.empty {{ text-align: center; color: var(--muted); padding: 48px 0; font-size: 13px; }}

    footer {{
        text-align: center;
        color: var(--muted);
        font-size: 11px;
        margin-top: 18px;
        letter-spacing: 0.3px;
    }}




    
        
            Live · Depth level 5
            {self.title}
            Auto-refreshes every 5s · {len(prices)} price levels tracked
        
    

    
        
            Total Bid Qty
            {total_bid:,}
        
        
            Total Ask Qty
            {total_ask:,}
        
        
            Net Imbalance
            {net_summary_sign}{total_net:,}
        
        
            Total Qty
            {(total_bid + total_ask):,}
        
        
            Last Refresh
            {now.strftime('%H:%M:%S')}
        
    

    
        {rows_joined}
            
            
                
                    Strike Price
                    Bid Qty (cum.)
                    Ask Qty (cum.)
                    Net (Bid − Ask)
                    Total Qty (Bid+Ask)
                    Last Updated
                
            
            
        
    
    Generated by PriceBookWriter · {now.strftime('%Y-%m-%d %H:%M:%S')}


"""

        with open(self.html_filepath, "w", encoding="utf-8") as f:
            f.write(html)